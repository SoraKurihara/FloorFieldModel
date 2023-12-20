import json
import os
import sys

import matplotlib.pyplot as plt
import numpy as np
import osmnx as ox
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from pyproj import Transformer
from scipy.ndimage import binary_fill_holes
from shapely.geometry import box
from tqdm import tqdm

cmap = plt.cm.colors.ListedColormap(["white", "red", "black", "green", "blue"])


def prompt_continue():
    while True:
        answer = input("上書きして実行を続けますか [y/n]? ").lower()
        if answer in ["y", "n"]:
            return answer == "y"
        else:
            print("無効な入力です。'y' または 'n' を入力してください。")


class CreateMap:
    def __init__(self, upperleft, lowerright, name="output"):
        self.min_lat = upperleft[0]
        self.min_lon = upperleft[1]
        self.max_lat = lowerright[0]
        self.max_lon = lowerright[1]
        self.name = name
        os.makedirs("map", exist_ok=True)
        if os.path.exists("Map.json"):
            with open("Map.json", "r") as file:
                self.existing_data = json.load(file)
        else:
            self.existing_data = {}

        if name in self.existing_data:
            print(f"Map.jsonに {name}キー はすでに存在しています．")
            if prompt_continue():
                print(f'"{name}"に上書きして実行を続けます。')
            else:
                print("実行を中止します。")
                sys.exit()

    def add_wall(self):
        Obstacles = np.where(self.grid >= 2)
        self.grid = self.grid[
            min(Obstacles[0]) - 2 : max(Obstacles[0]) + 2,
            min(Obstacles[1]) - 2 : max(Obstacles[1]) + 2,
        ]
        self.grid[0, :] = 2
        self.grid[-1, :] = 2
        self.grid[:, 0] = 2
        self.grid[:, -1] = 2

    def create_xlsx(self):
        # pandasのDataFrameに変換
        df = pd.DataFrame(self.grid)

        # 新しいExcelワークブックを作成
        wb = Workbook()
        ws = wb.active

        # DataFrameをExcelシートに変換
        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=False), 1
        ):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        max_row = df.shape[0]
        max_col = get_column_letter(df.shape[1])
        format_range = f"A1:{max_col}{max_row}"

        # 2の場合は黒で塗りつぶす（文字色も黒）
        black_fill = PatternFill(
            start_color="000001", end_color="000001", fill_type="solid"
        )
        black_font = Font(color="000000")
        ws.conditional_formatting.add(
            format_range,
            CellIsRule(
                operator="equal",
                formula=["2"],
                stopIfTrue=True,
                fill=black_fill,
                font=black_font,
            ),
        )
        # 5以上の場合も黒で塗りつぶす（文字色も黒）
        ws.conditional_formatting.add(
            format_range,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["5"],
                stopIfTrue=True,
                fill=black_fill,
                font=black_font,
            ),
        )

        # 3の場合は緑で塗りつぶす（文字色も緑）
        green_fill = PatternFill(
            start_color="00FF00", end_color="00FF00", fill_type="solid"
        )
        green_font = Font(color="00FF00")
        ws.conditional_formatting.add(
            format_range,
            CellIsRule(
                operator="equal",
                formula=["3"],
                stopIfTrue=True,
                fill=green_fill,
                font=green_font,
            ),
        )

        # 0の場合は白で塗りつぶす（文字色も白）
        white_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
        white_font = Font(color="FFFFFF")
        ws.conditional_formatting.add(
            format_range,
            CellIsRule(
                operator="equal",
                formula=["0"],
                stopIfTrue=True,
                fill=white_fill,
                font=white_font,
            ),
        )

        # 1の場合は赤で塗りつぶす（文字色も赤）
        red_fill = PatternFill(
            start_color="FF0000", end_color="FF0000", fill_type="solid"
        )
        red_font = Font(color="FF0000")
        ws.conditional_formatting.add(
            format_range,
            CellIsRule(
                operator="equal",
                formula=["1"],
                stopIfTrue=True,
                fill=red_fill,
                font=red_font,
            ),
        )

        # 4の場合は青で塗りつぶす（文字色も青）
        blue_fill = PatternFill(
            start_color="0000FF", end_color="0000FF", fill_type="solid"
        )
        blue_font = Font(color="0000FF")
        ws.conditional_formatting.add(
            format_range,
            CellIsRule(
                operator="equal",
                formula=["4"],
                stopIfTrue=True,
                fill=blue_fill,
                font=blue_font,
            ),
        )
        COLUMN_WIDTH_FACTOR = 6
        ROW_HEIGHT_FACTOR = 0.75

        # 正方形のサイズを15x15ピクセルに設定
        square_size = 15
        column_width = square_size / COLUMN_WIDTH_FACTOR  # 列の幅の設定値
        row_height = square_size * ROW_HEIGHT_FACTOR  # 行の高さの設定値

        # セルのサイズを正方形に設定
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = column_width

        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = row_height

        # ファイルに保存
        wb.save(rf"map/{self.name}.xlsx")
        print(rf"map/{self.name}.xlsx にxlsxが保存されているので出口などを書き換えてください．")

    def create(self, size=0.5):
        buildings = ox.features_from_bbox(
            north=self.max_lat,
            south=self.min_lat,
            west=self.min_lon,
            east=self.max_lon,
            tags={"building": True},
        )

        # Transformerを使用してWGS84の緯度経度座標系からUTM座標系への変換
        transformer = Transformer.from_crs(
            "EPSG:4326", "EPSG:32654", always_xy=True
        )
        utm_min_x, utm_min_y = transformer.transform(
            self.min_lon, self.min_lat
        )
        utm_max_x, utm_max_y = transformer.transform(
            self.max_lon, self.max_lat
        )

        # UTM座標でのバウンディングボックス
        x_min, y_min = utm_min_x, utm_min_y
        x_max, y_max = utm_max_x, utm_max_y

        # 建物のポリゴンをUTM座標に変換
        buildings_utm = buildings.to_crs(transformer.target_crs)

        cols = int(np.ceil((x_max - x_min) / size))
        rows = int(np.ceil((y_max - y_min) / size))
        self.grid = np.zeros((rows, cols), dtype=int)

        data = {}
        data["range"] = [
            self.min_lat,
            self.min_lon,
            self.max_lat,
            self.max_lon,
        ]

        for idx, building in tqdm(
            enumerate(buildings_utm.itertuples(), start=5)
        ):
            minx, miny, maxx, maxy = building.geometry.bounds
            # セルの座標を計算
            min_row = max(int((miny - y_min) / size), 0)
            max_row = min(int((maxy - y_min) / size), rows - 1)
            min_col = max(int((minx - x_min) / size), 0)
            max_col = min(int((maxx - x_min) / size), cols - 1)

            # R-treeインデックスを使用して、グリッドの範囲内で重なる建物のセルを特定
            for i in range(min_row, max_row + 1):
                for j in range(min_col, max_col + 1):
                    cell_box = box(
                        x_min + j * size,
                        y_min + i * size,
                        x_min + (j + 1) * size,
                        y_min + (i + 1) * size,
                    )
                    if building.geometry.intersects(cell_box):
                        self.grid[i, j] = 2

        building_mask = self.grid == 2
        filled_building_mask = binary_fill_holes(building_mask).astype(int)
        self.grid[filled_building_mask == 1] = 2

        self.existing_data[self.name] = data
        with open("Map.json", "w") as file:
            json.dump(self.existing_data, file, indent=4, ensure_ascii=False)
        print(self.grid)

        self.add_wall()
        self.create_xlsx()

        fig, ax = plt.subplots(figsize=(10, 10))
        im = ax.imshow(self.grid)
        fig.colorbar(im, ax=ax)
        plt.tight_layout()
        plt.show()


if __name__ == "__main__":
    min_lat = 34.87598184158452
    min_lon = 135.57329665538347
    max_lat = 34.879592178045755
    max_lon = 135.57754831825054

    upperleft = [min_lat, min_lon]
    lowerright = [max_lat, max_lon]
    cm = CreateMap(upperleft, lowerright, name="Takatsuki_")
    cm.create(size=0.5)
