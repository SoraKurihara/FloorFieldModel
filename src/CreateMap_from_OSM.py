import json
import os
import sys

import matplotlib.pyplot as plt
import numpy as np
import osmnx as ox
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from pyproj import Transformer
from scipy.ndimage import binary_fill_holes
from shapely.geometry import box
from tqdm import tqdm

cmap = plt.cm.colors.ListedColormap(["white", "red", "black", "green", "blue"])


def prompt_continue():
    while True:
        answer = input("上書きして実行を続けますか [y/n]? ").lower()
        if answer in ["y", "n"]:
            return answer == "y"
        else:
            print("無効な入力です。'y' または 'n' を入力してください。")


class CreateMap:
    def __init__(self, upperleft, lowerright, name="output"):
        self.min_lat = upperleft[0]
        self.min_lon = upperleft[1]
        self.max_lat = lowerright[0]
        self.max_lon = lowerright[1]
        self.name = name
        if not os.path.exists("Map"):
            os.makedirs("Map")
        if os.path.exists("Map.json"):
            with open("Map.json", "r") as file:
                existing_data = json.load(file)
        else:
            existing_data = {}
        if name in existing_data:
            print(f"Map.jsonに {name}キー はすでに存在しています．")
            if prompt_continue():
                print("上書きして実行を続けます。")
                existing_data[name] = upperleft + lowerright
            else:
                print("実行を中止します。")
                sys.exit()
        else:
            existing_data[name] = upperleft + lowerright

        with open("Map.json", "w") as file:
            json.dump(existing_data, file, indent=4)

    def create(self, size=0.5):
        buildings = ox.features_from_bbox(
            north=self.max_lat,
            south=self.min_lat,
            west=self.min_lon,
            east=self.max_lon,
            tags={"building": True},
        )

        # Transformerを使用してWGS84の緯度経度座標系からUTM座標系への変換
        transformer = Transformer.from_crs(
            "EPSG:4326", "EPSG:32654", always_xy=True
        )
        utm_min_x, utm_min_y = transformer.transform(
            self.min_lon, self.min_lat
        )
        utm_max_x, utm_max_y = transformer.transform(
            self.max_lon, self.max_lat
        )

        # UTM座標でのバウンディングボックス
        x_min, y_min = utm_min_x, utm_min_y
        x_max, y_max = utm_max_x, utm_max_y

        # グリッドの列数と行数を計算
        cols = int(np.ceil((x_max - x_min) / size))
        rows = int(np.ceil((y_max - y_min) / size))
        print(cols, rows)

        # グリッドを初期化
        grid = np.zeros((rows, cols), dtype=int)

        # GeoDataFrameをUTMに変換
        # GeoDataFrameをUTMに変換
        buildings_utm = buildings.to_crs(transformer.target_crs)

        # R-treeインデックスを作成
        sindex = buildings_utm.sindex

        # 各セルをチェックし、建物が含まれているかどうかを確認
        for i in tqdm(range(rows)):
            y_bottom = y_min + i * size
            y_top = y_bottom + size
            for j in tqdm(range(cols), leave=False):
                x_left = x_min + j * size
                x_right = x_left + size
                cell = box(x_left, y_bottom, x_right, y_top)
                # R-treeインデックスを使用して重なる建物のインデックスを取得
                possible_matches_index = list(sindex.intersection(cell.bounds))
                if possible_matches_index:
                    # 重なる建物があるかどうか確認
                    possible_matches = buildings_utm.iloc[
                        possible_matches_index
                    ]
                    precise_matches = possible_matches[
                        possible_matches.intersects(cell)
                    ]
                    if not precise_matches.empty:
                        grid[i, j] = 2  # 建物のセルを2でマーク

        # grid変数には、建物の位置を示す二次元配列が含まれています
        grid = np.flipud(grid)
        building_mask = grid == 2
        filled_building_mask = binary_fill_holes(building_mask).astype(int)
        grid[filled_building_mask == 1] = 2
        print(grid)
        np.save(rf"Map/{self.name}", grid)

        fig, ax = plt.subplots(figsize=(10, 10))
        ax.imshow(grid, cmap=cmap, vmin=0, vmax=3)
        plt.tight_layout()
        plt.show()

    def simple_wall(self):
        Wall = np.where(self.Map == 2)
        self.Map = self.Map[
            min(Wall[0]) - 2 : max(Wall[0]) + 2,
            min(Wall[1]) - 2 : max(Wall[1]) + 2,
        ]
        self.Map[0, :] = 2
        self.Map[-1, :] = 2
        self.Map[:, 0] = 2
        self.Map[:, -1] = 2
        return "SimpleWall"

    def center_exit(self):
        x_length = len(self.Map[0])
        x_center = int(x_length / 2)
        self.Map[0, (x_center - 5) : (x_center + 5)] = 3
        return "CenterExit"

    def add_wall(self, name=None):
        if name is None:
            name = self.name
        self.Map = np.load(rf"Map/{name}.npy")
        print(self.Map)

        wall = self.simple_wall()
        exit = self.center_exit()
        np.save(rf"Map/{name}_{wall}_{exit}", self.Map)

        fig, ax = plt.subplots(figsize=(10, 10))
        ax.imshow(self.Map, cmap=cmap, vmin=0, vmax=4)
        plt.tight_layout()
        plt.show()

    def create_xlsx(self, name=None):
        if name is None:
            name = self.name
        self.Map = np.load(rf"Map/{name}.npy")
        print(self.Map)
        # pandasのDataFrameに変換
        df = pd.DataFrame(self.Map)

        # 新しいExcelワークブックを作成
        wb = Workbook()
        ws = wb.active

        # DataFrameをExcelシートに変換
        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=False), 1
        ):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        max_row = df.shape[0]
        max_col = get_column_letter(df.shape[1])
        format_range = f"A1:{max_col}{max_row}"

        # 2の場合は黒で塗りつぶす（文字色も黒）
        black_fill = PatternFill(
            start_color="000001", end_color="000001", fill_type="solid"
        )
        black_font = Font(color="000000")
        ws.conditional_formatting.add(
            format_range,
            CellIsRule(
                operator="equal",
                formula=["2"],
                stopIfTrue=True,
                fill=black_fill,
                font=black_font,
            ),
        )

        # 3の場合は緑で塗りつぶす（文字色も緑）
        green_fill = PatternFill(
            start_color="00FF00", end_color="00FF00", fill_type="solid"
        )
        green_font = Font(color="00FF00")
        ws.conditional_formatting.add(
            format_range,
            CellIsRule(
                operator="equal",
                formula=["3"],
                stopIfTrue=True,
                fill=green_fill,
                font=green_font,
            ),
        )

        # 0の場合は白で塗りつぶす（文字色も白）
        white_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
        white_font = Font(color="FFFFFF")
        ws.conditional_formatting.add(
            format_range,
            CellIsRule(
                operator="equal",
                formula=["0"],
                stopIfTrue=True,
                fill=white_fill,
                font=white_font,
            ),
        )

        # 1の場合は赤で塗りつぶす（文字色も赤）
        red_fill = PatternFill(
            start_color="FF0000", end_color="FF0000", fill_type="solid"
        )
        red_font = Font(color="FF0000")
        ws.conditional_formatting.add(
            format_range,
            CellIsRule(
                operator="equal",
                formula=["1"],
                stopIfTrue=True,
                fill=red_fill,
                font=red_font,
            ),
        )

        # 4の場合は青で塗りつぶす（文字色も青）
        blue_fill = PatternFill(
            start_color="0000FF", end_color="0000FF", fill_type="solid"
        )
        blue_font = Font(color="0000FF")
        ws.conditional_formatting.add(
            format_range,
            CellIsRule(
                operator="equal",
                formula=["4"],
                stopIfTrue=True,
                fill=blue_fill,
                font=blue_font,
            ),
        )
        COLUMN_WIDTH_FACTOR = 6
        ROW_HEIGHT_FACTOR = 0.75

        # 正方形のサイズを15x15ピクセルに設定
        square_size = 15
        column_width = square_size / COLUMN_WIDTH_FACTOR  # 列の幅の設定値
        row_height = square_size * ROW_HEIGHT_FACTOR  # 行の高さの設定値

        # セルのサイズを正方形に設定
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = column_width

        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = row_height

        # ファイルに保存
        wb.save(rf"Map/{name}.xlsx")
        print(rf"Map/{name}.xlsx にxlsxが保存されているので出口などを書き換えてください．")
        sys.exit()


if __name__ == "__main__":
    min_lat = 34.87598184158452
    min_lon = 135.57329665538347
    max_lat = 34.879592178045755
    max_lon = 135.57754831825054

    upperleft = [min_lat, min_lon]
    lowerright = [max_lat, max_lon]
    cm = CreateMap(upperleft, lowerright, name="Takatsuki")
    cm.create(size=0.5)
    # cm.add_wall()
    cm.create_xlsx()
