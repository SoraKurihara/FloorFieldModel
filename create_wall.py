import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

filename = r"map/Takatsuki.xlsx"

Map = pd.read_excel(filename, header=None).fillna(0).values.astype(np.int8)

Obstacles = np.where(Map == 2)
Map = Map[
    min(Obstacles[0]) - 2 : max(Obstacles[0]) + 2,
    min(Obstacles[1]) - 2 : max(Obstacles[1]) + 2,
]
Map[0, :] = 2
Map[-1, :] = 2
Map[:, 0] = 2
Map[:, -1] = 2

print(Map)
# pandasのDataFrameに変換
df = pd.DataFrame(Map)

# 新しいExcelワークブックを作成
wb = Workbook()
ws = wb.active

# DataFrameをExcelシートに変換
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

max_row = df.shape[0]
max_col = get_column_letter(df.shape[1])
format_range = f"A1:{max_col}{max_row}"

# 2の場合は黒で塗りつぶす（文字色も黒）
black_fill = PatternFill(start_color="000001", end_color="000001", fill_type="solid")
black_font = Font(color="000000")
ws.conditional_formatting.add(
    format_range,
    CellIsRule(
        operator="equal",
        formula=["2"],
        stopIfTrue=True,
        fill=black_fill,
        font=black_font,
    ),
)

# 3の場合は緑で塗りつぶす（文字色も緑）
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
green_font = Font(color="00FF00")
ws.conditional_formatting.add(
    format_range,
    CellIsRule(
        operator="equal",
        formula=["3"],
        stopIfTrue=True,
        fill=green_fill,
        font=green_font,
    ),
)

# 0の場合は白で塗りつぶす（文字色も白）
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
white_font = Font(color="FFFFFF")
ws.conditional_formatting.add(
    format_range,
    CellIsRule(
        operator="equal",
        formula=["0"],
        stopIfTrue=True,
        fill=white_fill,
        font=white_font,
    ),
)

# 1の場合は赤で塗りつぶす（文字色も赤）
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
red_font = Font(color="FF0000")
ws.conditional_formatting.add(
    format_range,
    CellIsRule(
        operator="equal",
        formula=["1"],
        stopIfTrue=True,
        fill=red_fill,
        font=red_font,
    ),
)

# 4の場合は青で塗りつぶす（文字色も青）
blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
blue_font = Font(color="0000FF")
ws.conditional_formatting.add(
    format_range,
    CellIsRule(
        operator="equal",
        formula=["4"],
        stopIfTrue=True,
        fill=blue_fill,
        font=blue_font,
    ),
)
COLUMN_WIDTH_FACTOR = 6
ROW_HEIGHT_FACTOR = 0.75

# 正方形のサイズを15x15ピクセルに設定
square_size = 15
column_width = square_size / COLUMN_WIDTH_FACTOR  # 列の幅の設定値
row_height = square_size * ROW_HEIGHT_FACTOR  # 行の高さの設定値

# セルのサイズを正方形に設定
for col in range(1, ws.max_column + 1):
    ws.column_dimensions[get_column_letter(col)].width = column_width

for row in range(1, ws.max_row + 1):
    ws.row_dimensions[row].height = row_height

# ファイルに保存
wb.save(filename)
